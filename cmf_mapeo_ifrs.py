"""
cmf_mapeo_ifrs.py
=================
1. Lee el archivo Excel IFRS (estructura de cuentas contables)
2. Cruza los datos descargados (B1_historico.csv) con el mapeo IFRS
3. Genera el archivo consolidado con jerarquía IFRS completa
4. Crea el catálogo de bancos (código + nombre) como archivo Excel separado

Rutas esperadas (relativas al directorio del script, o ajusta DATA_DIR):
    data/
    ├── IFRS_mapeo.xlsx              ← tu archivo de mapeo (ponlo aquí)
    ├── consolidado/
    │   ├── B1_historico.csv
    │   ├── R1_historico.csv
    │   └── C1_historico.csv
    └── output/
        ├── B1_con_ifrs.csv          ← consolidado enriquecido con jerarquía IFRS
        ├── R1_con_ifrs.csv
        ├── C1_con_ifrs.csv
        ├── catalogo_bancos.xlsx     ← código + nombre de cada banco
        └── pivot_balance.xlsx       ← (opcional) tabla dinámica para exploración

Uso:
    python cmf_mapeo_ifrs.py
    python cmf_mapeo_ifrs.py --ifrs mi_mapeo.xlsx
    python cmf_mapeo_ifrs.py --solo-bancos   # solo regenera el catálogo de bancos
    python cmf_mapeo_ifrs.py --pivot         # genera además el pivot de balance
"""

import argparse
import logging
import sys
import pandas as pd
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN — ajusta estas rutas si mueves los archivos
# ─────────────────────────────────────────────────────────────────────────────
DATA_DIR    = Path("data")
CONSOL_DIR  = DATA_DIR / "consolidado"
OUTPUT_DIR  = DATA_DIR / "output"
IFRS_FILE   = DATA_DIR / "IFRS_mapeo.xlsx"   # nombre esperado del Excel IFRS

# Columnas del archivo IFRS (en el orden que aparecen en tu Excel)
IFRS_COLS = {
    "INFORME":      "informe",       # B / R / C
    "CODIGO_IFRS":  "cuenta",        # clave de join con los TXT descargados
    "NOMBRE":       "nombre_cuenta",
    "IFRS_NOMBRE":  "ifrs_nombre",
    "TIPO":         "tipo",
    "G0":           "g0",
    "G1":           "g1",
    "G2":           "g2",
    "G3":           "g3",
    "G4":           "g4",
    "G5":           "g5",
    "G6":           "g6",
    "G7":           "g7",
    "E1":           "e1",
    "E2":           "e2",
}

# Columnas de jerarquía que se agregan al consolidado
JERARQUIA_COLS = ["nombre_cuenta", "ifrs_nombre", "tipo", "g0", "g1", "g2",
                  "g3", "g4", "g5", "g6", "g7", "e1", "e2"]

# Nombres de las columnas numéricas en los TXT del B1
# Las 4 columnas corresponden a monedas: CLP, UF, MX1, MX2
COL_NOMBRES_B1 = {
    "col_1": "saldo_clp",
    "col_2": "saldo_uf",
    "col_3": "saldo_mx1",
    "col_4": "saldo_mx2",
}

COL_NOMBRES_R1 = {
    "col_1": "flujo_mes_actual",
}

COL_NOMBRES_C1 = {
    "col_1": "saldo_bruto",
    "col_2": "provisiones",
    "col_3": "saldo_neto",
    "col_4": "col_4",
}

COL_RENOMBRES = {"b1": COL_NOMBRES_B1, "r1": COL_NOMBRES_R1, "c1": COL_NOMBRES_C1}

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# 1. CARGA DEL MAPEO IFRS
# ─────────────────────────────────────────────────────────────────────────────
def cargar_ifrs(ruta: Path) -> pd.DataFrame:
    """
    Lee el Excel IFRS y devuelve un DataFrame normalizado listo para el join.
    La columna 'cuenta' queda como string sin espacios (para hacer match
    con la columna 'cuenta' de los TXT descargados).
    """
    if not ruta.exists():
        log.error(f"Archivo IFRS no encontrado: {ruta}")
        log.error("Coloca el Excel IFRS en la ruta indicada y vuelve a ejecutar.")
        sys.exit(1)

    log.info(f"Cargando mapeo IFRS: {ruta}")

    # Intentar leer la primera hoja; si el Excel tiene encabezados en fila 0
    df = pd.read_excel(ruta, dtype=str)

    # Renombrar columnas al estándar interno
    # (tolerante a mayúsculas/minúsculas y espacios extra en los nombres de columna)
    df.columns = [c.strip().upper() for c in df.columns]
    rename_map = {k.upper(): v for k, v in IFRS_COLS.items()}
    df = df.rename(columns=rename_map)

    # Verificar que existan las columnas mínimas
    cols_minimas = {"informe", "cuenta", "nombre_cuenta"}
    faltantes = cols_minimas - set(df.columns)
    if faltantes:
        log.error(f"Columnas faltantes en el Excel IFRS: {faltantes}")
        log.error(f"Columnas encontradas: {list(df.columns)}")
        sys.exit(1)

    # Limpiar la columna cuenta: quitar espacios, asegurar string
    df["cuenta"] = df["cuenta"].astype(str).str.strip()

    # Quitar filas sin código de cuenta
    df = df[df["cuenta"].notna() & (df["cuenta"] != "") & (df["cuenta"] != "nan")]

    # Asegurar que todas las columnas de jerarquía existan (rellenar con vacío si no)
    for col in JERARQUIA_COLS:
        if col not in df.columns:
            df[col] = ""

    log.info(f"  {len(df):,} cuentas cargadas del IFRS")
    log.info(f"  Informes disponibles: {sorted(df['informe'].dropna().unique())}")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# 2. CATÁLOGO DE BANCOS
# ─────────────────────────────────────────────────────────────────────────────
def generar_catalogo_bancos(df_hist: pd.DataFrame, ruta_salida: Path):
    """
    Extrae los pares únicos (banco_codigo, banco_nombre) del histórico
    y los guarda como Excel con formato.
    """
    catalogo = (
        df_hist[["banco_codigo", "banco_nombre"]]
        .drop_duplicates()
        .sort_values("banco_codigo")
        .reset_index(drop=True)
    )

    # Agregar columna de banco activo (para uso futuro)
    catalogo["activo"] = "Sí"

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        catalogo.to_excel(writer, sheet_name="Bancos", index=False)

        # Formato básico
        ws = writer.sheets["Bancos"]
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 12

        # Encabezado con color
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Filas alternas
        fill_par = PatternFill("solid", fgColor="D6E4F0")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = fill_par

    log.info(f"  ✓ Catálogo de bancos: {ruta_salida.name}  ({len(catalogo)} bancos)")
    return catalogo


# ─────────────────────────────────────────────────────────────────────────────
# 3. ENRIQUECIMIENTO CON IFRS
# ─────────────────────────────────────────────────────────────────────────────
def enriquecer_con_ifrs(
    df_hist: pd.DataFrame,
    df_ifrs: pd.DataFrame,
    tipo: str,
) -> pd.DataFrame:
    """
    Hace el join entre el histórico de datos CMF y el mapeo IFRS.
    Renombra las columnas numéricas según el tipo de informe.
    Reporta cuentas sin mapeo para revisión.
    """
    # Filtrar IFRS solo para este tipo de informe
    informe_letra = tipo[0].upper()   # "b1" → "B"
    ifrs_tipo = df_ifrs[df_ifrs["informe"].str.upper() == informe_letra].copy()

    # Asegurar que 'cuenta' sea string en ambos DataFrames
    df_hist = df_hist.copy()
    df_hist["cuenta"] = df_hist["cuenta"].astype(str).str.strip()

    # JOIN: left para mantener todos los datos aunque no haya mapeo IFRS
    columnas_ifrs = ["cuenta"] + JERARQUIA_COLS
    df_merged = df_hist.merge(
        ifrs_tipo[columnas_ifrs].drop_duplicates("cuenta"),
        on="cuenta",
        how="left",
    )

    # Renombrar columnas numéricas
    renombres = COL_RENOMBRES.get(tipo, {})
    renombres_validos = {k: v for k, v in renombres.items() if k in df_merged.columns}
    df_merged = df_merged.rename(columns=renombres_validos)

    # Para B1: limpiar comas y calcular totales por moneda
    if tipo == "b1":
        for col in ["saldo_clp", "saldo_uf", "saldo_mx1", "saldo_mx2"]:
            if col in df_merged.columns:
                df_merged[col] = (
                    df_merged[col].astype(str)
                    .str.replace(",", ".", regex=False)
                    .str.strip()
                )
                df_merged[col] = pd.to_numeric(df_merged[col], errors="coerce").fillna(0)

        # MX = MX1 + MX2
        if "saldo_mx1" in df_merged.columns and "saldo_mx2" in df_merged.columns:
            df_merged["saldo_mx"] = df_merged["saldo_mx1"] + df_merged["saldo_mx2"]
            df_merged.drop(columns=["saldo_mx1", "saldo_mx2"], inplace=True)

        # Total = CLP + UF + MX
        cols_sum = [c for c in ["saldo_clp", "saldo_uf", "saldo_mx"] if c in df_merged.columns]
        df_merged["saldo_total"] = df_merged[cols_sum].sum(axis=1)

    # Para R1: limpiar comas en flujo_mes_actual
    if tipo == "r1":
        if "flujo_mes_actual" in df_merged.columns:
            df_merged["flujo_mes_actual"] = (
                df_merged["flujo_mes_actual"].astype(str)
                .str.replace(",", ".", regex=False)
                .str.strip()
            )
            df_merged["flujo_mes_actual"] = pd.to_numeric(
                df_merged["flujo_mes_actual"], errors="coerce"
            ).fillna(0)

    # Reporte de cobertura
    sin_mapeo = df_merged[df_merged["nombre_cuenta"].isna()]["cuenta"].nunique()
    total_cuentas = df_merged["cuenta"].nunique()
    pct = (total_cuentas - sin_mapeo) / total_cuentas * 100 if total_cuentas else 0
    log.info(f"  {tipo.upper()} cobertura IFRS: {total_cuentas - sin_mapeo}/{total_cuentas} "
             f"cuentas mapeadas ({pct:.1f}%)")

    if sin_mapeo > 0:
        cuentas_sin = (
            df_merged[df_merged["nombre_cuenta"].isna()]["cuenta"]
            .value_counts()
            .head(10)
            .index.tolist()
        )
        log.warning(f"  Cuentas sin mapeo (top 10): {cuentas_sin}")

    return df_merged


# ─────────────────────────────────────────────────────────────────────────────
# 4. PIVOT OPCIONAL DE BALANCE
# ─────────────────────────────────────────────────────────────────────────────
def generar_pivot_balance(df_b1: pd.DataFrame, ruta_salida: Path):
    """
    Genera un pivot del balance: filas = cuenta/nombre, columnas = banco × período.
    Solo incluye cuentas con G4 definido (cuentas hoja con mapeo completo).
    """
    col_saldo = "saldo_mes_actual" if "saldo_mes_actual" in df_b1.columns else "col_1"

    df_filtrado = df_b1[df_b1["g4"].notna() & (df_b1["g4"] != "")].copy()
    if df_filtrado.empty:
        log.warning("  Sin datos con G4 para el pivot. Omitiendo.")
        return

    # Convertir saldo a numérico (los TXT traen 15 dígitos sin separador decimal)
    # La CMF usa miles sin decimales en pesos
    df_filtrado[col_saldo] = pd.to_numeric(df_filtrado[col_saldo], errors="coerce")

    pivot = df_filtrado.pivot_table(
        index=["cuenta", "nombre_cuenta", "g0", "g1", "g2", "g3", "g4"],
        columns=["periodo", "banco_codigo"],
        values=col_saldo,
        aggfunc="sum",
    )

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        pivot.to_excel(writer, sheet_name="Balance_Pivot")
        ws = writer.sheets["Balance_Pivot"]
        ws.freeze_panes = "H2"  # congelar columnas de jerarquía

    log.info(f"  ✓ Pivot balance: {ruta_salida.name}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Mapeo IFRS sobre datos CMF + catálogo de bancos"
    )
    parser.add_argument(
        "--ifrs", default=str(IFRS_FILE),
        help=f"Ruta al Excel IFRS (default: {IFRS_FILE})"
    )
    parser.add_argument(
        "--solo-bancos", action="store_true",
        help="Solo regenera el catálogo de bancos, sin rehacer el mapeo IFRS"
    )
    parser.add_argument(
        "--pivot", action="store_true",
        help="Genera además un pivot de balance en Excel"
    )
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Cargar IFRS ──────────────────────────────────────────────────────────
    if not args.solo_bancos:
        df_ifrs = cargar_ifrs(Path(args.ifrs))

    # ── Procesar cada tipo de informe ────────────────────────────────────────
    tipos_disponibles = []
    df_b1_enriquecido = None

    for tipo in ["b1", "r1", "c1"]:
        ruta_hist = CONSOL_DIR / f"{tipo.upper()}_historico.csv"

        if not ruta_hist.exists():
            log.info(f"No encontrado: {ruta_hist.name} — se omite.")
            continue

        log.info(f"Procesando {tipo.upper()}…")
        df_hist = pd.read_csv(ruta_hist, dtype=str, low_memory=False)

        # Normalizar banco_codigo a 3 dígitos con ceros a la izquierda
        df_hist["banco_codigo"] = (
            df_hist["banco_codigo"]
            .astype(str).str.strip()
            .apply(lambda x: str(int(x)).zfill(3) if x.isdigit() else x)
        )
        log.info(f"  {len(df_hist):,} filas cargadas")

        # ── Catálogo de bancos (se genera desde el primer tipo disponible) ──
        if not tipos_disponibles:
            ruta_cat = OUTPUT_DIR / "catalogo_bancos.xlsx"
            log.info("Generando catálogo de bancos…")
            generar_catalogo_bancos(df_hist, ruta_cat)

        tipos_disponibles.append(tipo)

        if args.solo_bancos:
            break   # solo necesitábamos el catálogo

        # ── Enriquecer con IFRS ──────────────────────────────────────────────
        df_enriquecido = enriquecer_con_ifrs(df_hist, df_ifrs, tipo)

        # Guardar CSV enriquecido
        ruta_out = OUTPUT_DIR / f"{tipo.upper()}_con_ifrs.csv"
        df_enriquecido.to_csv(ruta_out, index=False, encoding="utf-8-sig")
        log.info(f"  ✓ {ruta_out.name}  ({len(df_enriquecido):,} filas)")

        if tipo == "b1":
            df_b1_enriquecido = df_enriquecido

    # ── Pivot opcional ───────────────────────────────────────────────────────
    if args.pivot and df_b1_enriquecido is not None:
        log.info("Generando pivot de balance…")
        generar_pivot_balance(
            df_b1_enriquecido,
            OUTPUT_DIR / "pivot_balance.xlsx"
        )

    log.info("─" * 60)
    log.info("Proceso finalizado.")
    log.info(f"Archivos generados en: {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
